import imaplib
import email
import os
from io import BytesIO
import time
from openpyxl import load_workbook
import pywhatkit as kit
import datetime
from dotenv import load_dotenv

load_dotenv()

# Gmail settings
GMAIL_USERNAME = os.getenv('GMAIL_USERNAME')
GMAIL_PASSWORD = os.getenv('GMAIL_PASSWORD')
SEARCH_TEXT = os.getenv('SEARCH_TEXT')  

# Phone numbers
NUMBER1 = os.getenv('NUMBER1')  
NUMBER2 = os.getenv('NUMBER2')  # The phone number where you want to receive WhatsApp messages


# Connect to Gmail
def connect_to_gmail():
    try:
        mail = imaplib.IMAP4_SSL('imap.gmail.com')
        mail.login(GMAIL_USERNAME, GMAIL_PASSWORD)
        mail.select('inbox')
        print("Connected to Gmail")
        return mail
    except Exception as e:
        print(f"Error connecting to Gmail: {str(e)}")
        return None

# Send WhatsApp message
def send_whatsapp_message(number, subject):
    message = f"New Email Subject: {subject}"
    kit.sendwhatmsg(number, message, datetime.datetime.now().hour,datetime.datetime.now().minute+1, 15, True, 1)
    print(f"WhatsApp message sent to {number}")

# Search for SEARCH_TEXT in email attachments
def search_text_in_attachments(msg):
    for part in msg.walk():
        if part.get_content_maintype() == 'multipart':
            continue
        if part.get_content_maintype() == 'text':
            continue
        filename = part.get_filename()
        if filename:
            if filename.lower().endswith('.xlsx') or filename.lower().endswith('.xls'):
                # Check if the attachment is an XLSX or XLS file
                attachment_content = part.get_payload(decode=True)
                with BytesIO(attachment_content) as excel_buffer:
                    wb = load_workbook(excel_buffer, data_only=True)
                    for sheet in wb:
                        for row in sheet.iter_rows(values_only=True):
                            for cell_value in row:
                                if SEARCH_TEXT.lower() in str(cell_value).lower():
                                    return True
            else:
                attachment = part.get_payload(decode=True).decode('utf-8', 'ignore')
                if SEARCH_TEXT.lower() in attachment.lower():
                    return True
    return False

# Monitor Gmail for the latest email
def monitor_gmail():
    mail = connect_to_gmail()
    prev_subject = ""
    while True:
        if mail:
            # Search for the latest email
            status, email_data = mail.uid('search', None, 'ALL')
            if status == 'OK':
                email_uids = email_data[0].split()
                if email_uids:
                    latest_email_uid = email_uids[-1]  # Get the latest email UID
                    status, email_data = mail.uid('fetch', latest_email_uid, '(RFC822)')
                    if status == 'OK':
                        msg = email.message_from_bytes(email_data[0][1])

                        # Initialize an empty variable to store the email body text
                        email_body = ""

                        # Iterate through the message parts and concatenate the text
                        for part in msg.walk():
                            if part.get_content_type() == "text/plain":
                                email_body += part.get_payload(decode=True).decode("utf-8", "ignore")

                        print(f"Checking email body:\n{email_body}")

                        if SEARCH_TEXT.lower() in email_body.lower() and msg['subject'].lower() != prev_subject:
                            prev_subject = msg['subject'].lower()
                            print(f"CDC Mail: {msg['subject']}")
                            send_whatsapp_message(NUMBER2, msg['subject'])

                        # Check attachments for SEARCH_TEXT and send a WhatsApp message to NUMBER2
                        if search_text_in_attachments(msg) and msg['subject'].lower() != prev_subject:
                            prev_subject = msg['subject'].lower()
                            print(f"CDC Mail: {msg['subject']}")
                            send_whatsapp_message(NUMBER2, msg['subject'])

            # mail.logout()
        print("Sleeping for 10 seconds...")
        time.sleep(10)

# Wait for 10 seconds before checking for the latest email again

monitor_gmail()
